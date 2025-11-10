"""Split each data row of an Excel file into its own workbook."""
from openpyxl import load_workbook, Workbook


def split_rows(input_file: str):
    wb = load_workbook(input_file)
    ws = wb.active

    header = [cell.value for cell in ws[1]]

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.append(header)
        out_ws.append(list(row))
        out_name = f"Pattern{idx}.xlsx"
        out_wb.save(out_name)
        print(f"Created {out_name}")


if __name__ == "__main__":
    split_rows("input.xlsx")
